import os
import sys
from PIL import Image
from tqdm import tqdm
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# グリッドのサイズ
GRID_SIZE = 10

# ExcelのPixelサイズ変換係数
COL_PP = 0.12438
ROW_PP = 0.75

def PrintUsage():
    print('Usage: python this.py <img_file>')

if __name__ == '__main__':
    argl = len(sys.argv)
    if argl != 2 :
        PrintUsage()
        sys.exit()

    file_name = sys.argv[1]
    img = Image.open(file_name).convert('RGBA')
    width, height = img.size
    imgdata = img.getdata()

    #Workbook作成
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Excel Art'

    for y in tqdm(range(height)):
        y_offset = y * width
        for x in range(width):
            #画像から色データ取得
            r,g,b,a = imgdata[y_offset+x]
            if a > 128: #透明はスキップ
                #セルに設定
                color_ff = '%02X%02X%02X' % (r,g,b)
                sheet.cell(row=y+1, column=x+1).fill = PatternFill(patternType='solid', fgColor=color_ff)

            #グリッドサイズを正方形に設定
            sheet.column_dimensions[get_column_letter(x+1)].width = GRID_SIZE * COL_PP
            sheet.row_dimensions[y+1].height = GRID_SIZE * ROW_PP

    #保存
    file_name_body, ext = os.path.splitext(os.path.basename(file_name))
    wb.save(file_name_body+'.xlsx')
